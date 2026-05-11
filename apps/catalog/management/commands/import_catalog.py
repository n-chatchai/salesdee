"""Import products from an .xlsx file (header row required), mapped by column name.

    uv run python manage.py import_catalog path/to/catalog.xlsx --tenant <slug>

Recognised headers (case-insensitive, Thai or English): code/รหัส/sku, name/ชื่อ, name_en, unit/หน่วย,
price/ราคา, cost/ต้นทุน, tax/ภาษี, category/หมวด, material/วัสดุ, finish/สี, color_code/รหัสสี,
hardware_brand, standard/มอก, width_mm/กว้าง, depth_mm/ลึก, height_mm/สูง, tags/แท็ก.
Rows with no name are skipped. Existing products are matched by `code` (within the tenant) and updated.

This is the basic importer; a web upload wizard (column mapping, preview/confirm, error report) is
a later task — see NEXT_STEPS.
"""

from __future__ import annotations

import contextlib
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError, CommandParser

from apps.catalog.models import Product, ProductCategory, TaxType
from apps.core.current_tenant import tenant_context
from apps.tenants.models import Tenant

_TAX_MAP = {
    "": TaxType.VAT7,
    "7": TaxType.VAT7,
    "vat7": TaxType.VAT7,
    "vat 7%": TaxType.VAT7,
    "vat7%": TaxType.VAT7,
    "0": TaxType.VAT0,
    "vat0": TaxType.VAT0,
    "vat 0%": TaxType.VAT0,
    "exempt": TaxType.EXEMPT,
    "ยกเว้น": TaxType.EXEMPT,
    "none": TaxType.NONE,
    "no": TaxType.NONE,
    "ไม่คิด": TaxType.NONE,
}
_DECIMAL_FIELDS = {"default_price", "cost"}
_INT_FIELDS = {"width_mm", "depth_mm", "height_mm"}
# header (lowercased) -> Product field name ("_category" handled specially)
_FIELD_MAP = {
    "code": "code", "รหัส": "code", "sku": "code",
    "name": "name", "ชื่อ": "name", "ชื่อสินค้า": "name",
    "name_en": "name_en", "ชื่ออังกฤษ": "name_en",
    "description": "description", "รายละเอียด": "description",
    "unit": "unit", "หน่วย": "unit",
    "price": "default_price", "ราคา": "default_price", "default_price": "default_price",
    "cost": "cost", "ต้นทุน": "cost",
    "tax": "tax_type", "tax_type": "tax_type", "ภาษี": "tax_type",
    "category": "_category", "หมวด": "_category",
    "material": "material", "วัสดุ": "material",
    "finish": "finish", "สี": "finish",
    "color_code": "color_code", "รหัสสี": "color_code",
    "hardware_brand": "hardware_brand", "ยี่ห้ออุปกรณ์": "hardware_brand",
    "standard": "standard", "มอก": "standard",
    "width_mm": "width_mm", "กว้าง": "width_mm",
    "depth_mm": "depth_mm", "ลึก": "depth_mm",
    "height_mm": "height_mm", "สูง": "height_mm",
    "tags": "tags", "แท็ก": "tags",
}  # fmt: skip


class Command(BaseCommand):
    help = "Import products from an .xlsx file. Use --tenant <slug>."

    def add_arguments(self, parser: CommandParser) -> None:
        parser.add_argument("path")
        parser.add_argument("--tenant", dest="tenant_slug", required=True)

    def handle(self, *args, **opts) -> None:
        from openpyxl import load_workbook

        try:
            tenant = Tenant.objects.get(slug=opts["tenant_slug"])
        except Tenant.DoesNotExist as exc:
            raise CommandError(f"ไม่พบ tenant slug={opts['tenant_slug']!r}") from exc

        wb = load_workbook(opts["path"], read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise CommandError("ไฟล์ไม่มี worksheet")
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError("ไฟล์ว่าง (ไม่มีแถวหัวตาราง)")
        cols = [_FIELD_MAP.get(str(h).strip().lower()) if h else None for h in header]

        created = updated = skipped = 0
        with tenant_context(tenant):
            cat_cache: dict[str, ProductCategory] = {}
            for row in rows:
                data: dict = {}
                category: ProductCategory | None = None
                for col, value in zip(cols, row, strict=False):
                    if col is None or value in (None, ""):
                        continue
                    if col == "_category":
                        cname = str(value).strip()
                        category = cat_cache.get(cname)
                        if category is None:
                            category, _ = ProductCategory.objects.get_or_create(name=cname)
                            cat_cache[cname] = category
                    elif col == "tax_type":
                        data["tax_type"] = _TAX_MAP.get(str(value).strip().lower(), TaxType.VAT7)
                    elif col in _DECIMAL_FIELDS:
                        with contextlib.suppress(InvalidOperation, ValueError):
                            data[col] = Decimal(str(value).replace(",", "").strip())
                    elif col in _INT_FIELDS:
                        with contextlib.suppress(ValueError):
                            data[col] = int(float(str(value).replace(",", "").strip()))
                    else:
                        data[col] = str(value).strip()
                if not data.get("name"):
                    skipped += 1
                    continue
                if category is not None:
                    data["category"] = category
                code = data.pop("code", "")
                if code:
                    _, was_created = Product.objects.update_or_create(code=code, defaults=data)
                else:
                    Product.objects.create(**data)
                    was_created = True
                created += int(was_created)
                updated += int(not was_created)
        self.stdout.write(
            self.style.SUCCESS(f"นำเข้าเสร็จ: สร้าง {created}, อัปเดต {updated}, ข้าม {skipped}")
        )
